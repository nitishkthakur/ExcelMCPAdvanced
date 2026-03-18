"""
Excel reading utilities.
- Values are read via python-calamine (fast, native).
- Formulas are read via openpyxl (for hybrid content mode only).
- Sheet size uses direct ZIP/XML parsing for .xlsx (O(1) on file size).
"""

from __future__ import annotations
import re
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET


def get_sheet_names(file_path: str) -> list[str]:
    """Return all sheet names in the workbook (fast, calamine)."""
    from python_calamine import CalamineWorkbook
    wb = CalamineWorkbook.from_path(file_path)
    return list(wb.sheet_names)


def read_sheet_values(file_path: str, sheet_name: str) -> list[list[Any]]:
    """
    Read all cell values for the given sheet as a 2-D list (rows × cols).
    Uses calamine for maximum speed. Returns native Python types.
    Trailing empty rows are stripped.
    """
    from python_calamine import CalamineWorkbook
    wb = CalamineWorkbook.from_path(file_path)
    sheet = wb.get_sheet_by_name(sheet_name)
    rows = sheet.to_python(skip_empty_area=False)

    # Strip trailing all-empty rows
    while rows and all(v is None or v == "" for v in rows[-1]):
        rows.pop()
    return rows


def get_sheet_size(file_path: str, sheet_name: str) -> tuple[int, int, str]:
    """
    Return (n_rows, n_cols, method) for the sheet.

    Strategy:
      .xlsx / .xlsm — parse <dimension> tag directly from the ZIP file.
        Reads workbook.xml (~2 KB) + the first 8 KB of the sheet XML.
        O(1) on file size: a 200 MB sheet takes the same time as a 10 KB sheet.
      All other formats (.xls, .xlsb, .ods) OR xlsx with a missing/malformed
        dimension tag — iterate with calamine, counting rows and max column width.
        O(n rows) but memory-light: rows are never fully materialised.

    Returns:
      (rows, cols, method)  where method is 'xml_dimension_tag' or 'calamine_iteration'.
    """
    path = Path(file_path).expanduser().resolve()
    suffix = path.suffix.lower()

    if suffix in (".xlsx", ".xlsm", ".xlam"):
        result = _xlsx_dimension_from_zip(str(path), sheet_name)
        if result is not None:
            rows, cols = result
            return rows, cols, "xml_dimension_tag"

    rows, cols = _size_via_calamine(str(path), sheet_name)
    return rows, cols, "calamine_iteration"


def _xlsx_dimension_from_zip(path: str, sheet_name: str) -> tuple[int, int] | None:
    """
    Parse <dimension ref="A1:E15000"/> directly from the xlsx ZIP.

    Reads:
      • xl/workbook.xml          — to map sheet name → relationship ID
      • xl/_rels/workbook.xml.rels — to map relationship ID → sheet XML path
      • First 8 KB of the sheet XML — to find the <dimension> element

    Returns (n_rows, n_cols) or None if the element is absent or unparseable.
    """
    NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

    try:
        with zipfile.ZipFile(path) as zf:
            wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
            rels_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))

            rid_to_target: dict[str, str] = {
                r.get("Id"): r.get("Target") for r in rels_root
            }

            sheet_rid: str | None = None
            for sh in wb_root.iter(f"{{{NS_MAIN}}}sheet"):
                if sh.get("name") == sheet_name:
                    sheet_rid = sh.get(f"{{{NS_R}}}id")
                    break

            if not sheet_rid or sheet_rid not in rid_to_target:
                return None

            target = rid_to_target[sheet_rid]
            # target is 'worksheets/sheet1.xml' or '/xl/worksheets/sheet1.xml'
            sheet_path = (
                target.lstrip("/") if target.startswith("/") else f"xl/{target}"
            )

            # Read only the first 8 KB of the sheet XML
            with zf.open(sheet_path) as f:
                chunk = f.read(8192).decode("utf-8", errors="replace")

        m = re.search(r'<dimension\s[^>]*ref="([^"]+)"', chunk)
        if not m:
            return None

        ref = m.group(1)
        if ":" not in ref:
            # Single-cell sheet (e.g. ref="A1")
            return 1, 1

        tl, br = ref.split(":", 1)
        from .patches import parse_a1
        r1, c1 = parse_a1(tl)
        r2, c2 = parse_a1(br)
        return r2 - r1 + 1, c2 - c1 + 1

    except Exception:
        return None


def _size_via_calamine(path: str, sheet_name: str) -> tuple[int, int]:
    """
    Count rows and max column width by iterating with calamine.
    Rows are consumed one at a time — the full sheet is never held in memory.
    """
    from python_calamine import CalamineWorkbook
    wb = CalamineWorkbook.from_path(path)
    sheet = wb.get_sheet_by_name(sheet_name)
    n_rows = 0
    n_cols = 0
    for row in sheet.iter_rows():
        n_rows += 1
        if len(row) > n_cols:
            n_cols = len(row)
    return n_rows, n_cols


def read_sheet_formulas(file_path: str, sheet_name: str) -> dict[tuple[int, int], str]:
    """
    Return a dict of {(row_0idx, col_0idx): formula_string} for cells that
    contain Excel formulas in the given sheet.
    Uses openpyxl (slower; call only for hybrid content mode).
    """
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=False)
    ws = wb[sheet_name]
    formulas: dict[tuple[int, int], str] = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                # openpyxl uses 1-based row/col; convert to 0-based
                formulas[(cell.row - 1, cell.column - 1)] = cell.value
    wb.close()
    return formulas
