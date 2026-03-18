"""
Excel writing from markdown table input.
"""

from __future__ import annotations
import re
from typing import Any


def parse_markdown_table(md: str) -> list[list[str]]:
    """
    Parse a markdown table string into a 2-D list of strings.
    Separator rows (|---|---| etc.) are skipped.
    Pipe characters within cell values must be escaped as \\|.
    """
    rows: list[list[str]] = []
    for line in md.strip().splitlines():
        line = line.strip()
        if not line.startswith("|"):
            continue
        # Check if this is a separator row
        inner = line.strip("|")
        if re.match(r"^[\s\-\:]+(\|[\s\-\:]+)*$", inner):
            continue
        # Split on unescaped pipes
        raw_cells = re.split(r"(?<!\\)\|", line)
        # First and last are empty due to leading/trailing |
        cells = [c.strip().replace("\\|", "|") for c in raw_cells[1:-1]]
        rows.append(cells)
    return rows


def coerce_cell(value: str) -> Any:
    """
    Try to coerce a string cell value to a native Python type:
    - Formula strings (starting with =) kept as-is
    - Integers
    - Floats
    - Booleans (TRUE/FALSE)
    - Empty string → None
    - Otherwise keep as str
    """
    if not value or value.strip() == "":
        return None
    s = value.strip()
    if s.startswith("="):
        return s  # formula
    if s.upper() == "TRUE":
        return True
    if s.upper() == "FALSE":
        return False
    # Strip common formatting: $, %, commas
    clean = s.replace(",", "").replace("$", "").replace("%", "")
    try:
        iv = int(clean)
        return iv
    except ValueError:
        pass
    try:
        fv = float(clean)
        # If original had %, divide by 100
        if s.endswith("%"):
            return fv / 100
        return fv
    except ValueError:
        pass
    return s


def write_excel(file_path: str, sheets_data: dict[str, str]) -> None:
    """
    Write an Excel workbook to file_path.

    sheets_data: {sheet_name: markdown_table_string}
    Each markdown table's first row becomes the column header.
    Cell values are coerced: formulas written as formulas, numbers as numbers.
    """
    import openpyxl
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for sheet_name, md in sheets_data.items():
        ws = wb.create_sheet(title=sheet_name)
        rows = parse_markdown_table(md)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, cell_str in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=coerce_cell(cell_str))

    wb.save(file_path)
