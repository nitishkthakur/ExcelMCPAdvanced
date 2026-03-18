"""
Creates a multi-sheet test Excel file for testing the Excel MCP Server.
Run: python tests/create_test_excel.py
"""

import openpyxl
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import random
import os

random.seed(42)

OUTPUT_PATH = "tests/test_workbook.xlsx"


def create_test_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Sales (100 rows of order data) ───────────────────────────────
    ws_sales = wb.create_sheet("Sales")
    headers = ["Order ID", "Date", "Customer", "Product", "Amount"]
    ws_sales.append(headers)
    customers = ["Apex Corp", "Beta LLC", "Charlie Inc", "Delta Co", "Epsilon Ltd"]
    products = ["Widget A", "Widget B", "Widget C", "Gadget X", "Gadget Y"]
    start_date = date(2024, 1, 1)
    for i in range(1, 101):
        row = [
            10000 + i,
            (start_date + timedelta(days=i - 1)).isoformat(),
            random.choice(customers),
            random.choice(products),
            round(random.uniform(50, 500), 2),
        ]
        ws_sales.append(row)

    # ── Sheet 2: Forecast (5 rows × 50 cols, starting at B5) ──────────────────
    ws_fc = wb.create_sheet("Forecast")
    # Leave rows 1-4 and col A empty to test non-A1 patch detection
    metrics = ["Revenue", "COGS", "Gross Margin", "Operating Exp", "Net Income"]
    # Write header at B5 (row 5, col 2)
    ws_fc.cell(row=5, column=2, value="Metric")
    for j in range(1, 51):
        ws_fc.cell(row=5, column=j + 2, value=f"Month_{j:02d}")
    for i, metric in enumerate(metrics, start=1):
        ws_fc.cell(row=5 + i, column=2, value=metric)
        for j in range(1, 51):
            ws_fc.cell(row=5 + i, column=j + 2, value=round(random.uniform(1000, 50000), 2))

    # ── Sheet 3: Config (small 5-row table) ───────────────────────────────────
    ws_cfg = wb.create_sheet("Config")
    ws_cfg.append(["Setting", "Value"])
    ws_cfg.append(["tax_rate", 0.2])
    ws_cfg.append(["discount", 0.05])
    ws_cfg.append(["currency", "USD"])
    ws_cfg.append(["version", "2.1.0"])

    # ── Sheet 4: Formulas ─────────────────────────────────────────────────────
    ws_f = wb.create_sheet("Formulas")
    ws_f.append(["Item", "Price", "Qty", "Total"])
    ws_f.append(["Apple", 1.5, 10, "=B2*C2"])
    ws_f.append(["Banana", 0.5, 20, "=B3*C3"])
    ws_f.append(["Cherry", 3.0, 5, "=B4*C4"])
    ws_f.append(["", "", "Grand Total", "=SUM(D2:D4)"])

    # ── Sheet 5: Sparse (isolated patches) ────────────────────────────────────
    ws_sp = wb.create_sheet("Sparse")
    # Patch 1: top-left
    ws_sp.cell(row=1, column=1, value="Key")
    ws_sp.cell(row=1, column=2, value="Value")
    ws_sp.cell(row=2, column=1, value="alpha")
    ws_sp.cell(row=2, column=2, value=100)
    ws_sp.cell(row=3, column=1, value="beta")
    ws_sp.cell(row=3, column=2, value=200)
    # Empty rows 4-7
    # Patch 2: bottom area
    ws_sp.cell(row=8, column=5, value="Name")
    ws_sp.cell(row=8, column=6, value="Score")
    ws_sp.cell(row=9, column=5, value="Alice")
    ws_sp.cell(row=9, column=6, value=95)
    ws_sp.cell(row=10, column=5, value="Bob")
    ws_sp.cell(row=10, column=6, value=87)

    os.makedirs("tests", exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Test workbook created: {OUTPUT_PATH}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    create_test_workbook()
